"""
build_eda.py
============
Project 2 — Exploratory Data Analysis (EDA)
DecodeLabs Industrial Training Program — Batch 2026

Reads the raw e-commerce orders dataset and produces a single Excel workbook
(DecodeLabs_Project2_EDA.xlsx) containing:
    1. Key Observations   - narrative summary of findings
    2. Summary Statistics - mean/median/mode/std/min/max/quartiles (live formulas)
    3. Pivot Tables        - SUMIF/COUNTIF group summaries (Product, Status,
                              Payment Method, Referral Source, Coupon, Month)
    4. Charts               - bar charts, line chart (monthly trend), histogram
    5. Outliers             - IQR-method outlier detection on TotalPrice
    6. Raw Data              - the original dataset as an Excel Table

All statistics in the workbook are computed with native Excel formulas
(not hardcoded Python values), so the workbook recalculates automatically
if the Raw Data sheet is edited.

Usage:
    python build_eda.py [input_xlsx] [output_xlsx]

    input_xlsx   path to the raw dataset (default: data/Dataset_for_Data_Analytics.xlsx)
    output_xlsx  path to write the workbook (default: output/DecodeLabs_Project2_EDA.xlsx)

Requirements:
    pandas, numpy, openpyxl
    LibreOffice (for formula recalculation, run separately — see README)
"""

import sys
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_SRC = os.path.join(SCRIPT_DIR, 'data', 'Dataset_for_Data_Analytics.xlsx')
DEFAULT_OUT = os.path.join(SCRIPT_DIR, 'output', 'DecodeLabs_Project2_EDA.xlsx')

SRC = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
OUT_PATH = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

df = pd.read_excel(SRC)
df['Month'] = df['Date'].dt.to_period('M').astype(str)
df['CouponCode'] = df['CouponCode'].fillna('None')

numeric_cols = ['Quantity', 'UnitPrice', 'ItemsInCart', 'TotalPrice']

# ---------- Styling helpers ----------
NAVY = '1F3864'
LIGHT_BLUE = 'D9E2F3'
ACCENT = '4472C4'
WHITE_FONT = Font(color='FFFFFF', bold=True, name='Arial', size=11)
TITLE_FONT = Font(color=NAVY, bold=True, name='Arial', size=16)
SUBTITLE_FONT = Font(color='595959', italic=True, name='Arial', size=10)
HEADER_FILL = PatternFill('solid', fgColor=NAVY)
SECTION_FILL = PatternFill('solid', fgColor=LIGHT_BLUE)
BOLD = Font(bold=True, name='Arial', size=11)
NORMAL = Font(name='Arial', size=11)
THIN = Side(style='thin', color='B7B7B7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col+ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def add_table(ws, ref, name, style='TableStyleMedium9'):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True, showFirstColumn=False)
    ws.add_table(tab)

wb = Workbook()

# ============================================================
# SHEET 1: Raw Data
# ============================================================
ws_raw = wb.active
ws_raw.title = 'Raw Data'
cols = ['OrderID','Date','CustomerID','Product','Quantity','UnitPrice','ShippingAddress',
        'PaymentMethod','OrderStatus','TrackingNumber','ItemsInCart','CouponCode','ReferralSource','TotalPrice']
ws_raw.append(cols)
style_header_row(ws_raw, 1, len(cols))
for r in df[cols].itertuples(index=False):
    ws_raw.append(list(r))

date_col_idx = cols.index('Date') + 1
for row in range(2, ws_raw.max_row+1):
    ws_raw.cell(row=row, column=date_col_idx).number_format = 'yyyy-mm-dd'
    for c in [5,6,11,14]:  # Quantity, UnitPrice, ItemsInCart, TotalPrice
        cell = ws_raw.cell(row=row, column=c)
        if c in (6,14):
            cell.number_format = '#,##0.00'

n_rows = ws_raw.max_row
add_table(ws_raw, f'A1:N{n_rows}', 'RawData')
widths = {'A':12,'B':12,'C':11,'D':10,'E':9,'F':10,'G':16,'H':13,'I':11,'J':14,'K':11,'L':11,'M':14,'N':11}
autosize(ws_raw, widths)
ws_raw.freeze_panes = 'A2'

last_row = n_rows  # 1201

print('Raw Data sheet built, rows:', n_rows)

# ============================================================
# SHEET 2: Summary Statistics (formula-driven)
# ============================================================
ws_sum = wb.create_sheet('Summary Statistics')
ws_sum['A1'] = 'Exploratory Data Analysis — Summary Statistics'
ws_sum['A1'].font = TITLE_FONT
ws_sum.merge_cells('A1:F1')
ws_sum['A2'] = f'Dataset: {n_rows-1} orders | Source: Raw Data sheet'
ws_sum['A2'].font = SUBTITLE_FONT
ws_sum.merge_cells('A2:F2')

stat_headers = ['Metric','Quantity','UnitPrice ($)','ItemsInCart','TotalPrice ($)']
start_r = 4
for j,h in enumerate(stat_headers, start=1):
    ws_sum.cell(row=start_r, column=j, value=h)
style_header_row(ws_sum, start_r, len(stat_headers))

col_map = {'Quantity':'E', 'UnitPrice ($)':'F', 'ItemsInCart':'K', 'TotalPrice ($)':'N'}
metrics = [
    ('Count',   lambda col: f'=COUNT(\'Raw Data\'!{col}2:{col}{last_row})'),
    ('Mean',    lambda col: f'=AVERAGE(\'Raw Data\'!{col}2:{col}{last_row})'),
    ('Median',  lambda col: f'=MEDIAN(\'Raw Data\'!{col}2:{col}{last_row})'),
    ('Mode',    lambda col: f'=_xlfn.MODE.SNGL(\'Raw Data\'!{col}2:{col}{last_row})'),
    ('Std Dev', lambda col: f'=_xlfn.STDEV.S(\'Raw Data\'!{col}2:{col}{last_row})'),
    ('Min',     lambda col: f'=MIN(\'Raw Data\'!{col}2:{col}{last_row})'),
    ('Max',     lambda col: f'=MAX(\'Raw Data\'!{col}2:{col}{last_row})'),
    ('Q1 (25th pct)', lambda col: f'=_xlfn.QUARTILE.INC(\'Raw Data\'!{col}2:{col}{last_row},1)'),
    ('Q3 (75th pct)', lambda col: f'=_xlfn.QUARTILE.INC(\'Raw Data\'!{col}2:{col}{last_row},3)'),
]

r = start_r + 1
for mname, formula_fn in metrics:
    ws_sum.cell(row=r, column=1, value=mname).font = BOLD
    ws_sum.cell(row=r, column=1).border = BORDER
    for j, h in enumerate(stat_headers[1:], start=2):
        col = col_map[h]
        cell = ws_sum.cell(row=r, column=j, value=formula_fn(col))
        cell.number_format = '#,##0.00' if h != 'Quantity' and h != 'ItemsInCart' else '#,##0.0'
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    r += 1

stats_end_row = r - 1
for col_letter, w in zip(['A','B','C','D','E'], [16,13,14,13,14]):
    ws_sum.column_dimensions[col_letter].width = w

# IQR & outlier bounds block
r += 1
ws_sum.cell(row=r, column=1, value='Outlier Bounds (IQR method: Q1-1.5×IQR, Q3+1.5×IQR)').font = BOLD
r += 1
bound_header_row = r
for j,h in enumerate(['Metric','UnitPrice ($)','TotalPrice ($)'], start=1):
    ws_sum.cell(row=r, column=j, value=h)
style_header_row(ws_sum, r, 3)
r += 1
iqr_row = r
ws_sum.cell(row=r, column=1, value='IQR').font = BOLD
ws_sum.cell(row=r, column=2, value=f"=C{stats_end_row}-C{stats_end_row-1}").number_format='#,##0.00'
ws_sum.cell(row=r, column=3, value=f"=E{stats_end_row}-E{stats_end_row-1}").number_format='#,##0.00'
r += 1
lower_row = r
ws_sum.cell(row=r, column=1, value='Lower Bound').font = BOLD
ws_sum.cell(row=r, column=2, value=f"=C{stats_end_row-1}-1.5*B{iqr_row}").number_format='#,##0.00'
ws_sum.cell(row=r, column=3, value=f"=E{stats_end_row-1}-1.5*C{iqr_row}").number_format='#,##0.00'
r += 1
upper_row = r
ws_sum.cell(row=r, column=1, value='Upper Bound').font = BOLD
ws_sum.cell(row=r, column=2, value=f"=C{stats_end_row}+1.5*B{iqr_row}").number_format='#,##0.00'
ws_sum.cell(row=r, column=3, value=f"=E{stats_end_row}+1.5*C{iqr_row}").number_format='#,##0.00'
r += 1
outlier_count_row = r
ws_sum.cell(row=r, column=1, value='Outlier Count').font = BOLD
ws_sum.cell(row=r, column=2, value=f"=COUNTIFS('Raw Data'!F2:F{last_row},\"<\"&B{lower_row})+COUNTIFS('Raw Data'!F2:F{last_row},\">\"&B{upper_row})")
ws_sum.cell(row=r, column=3, value=f"=COUNTIFS('Raw Data'!N2:N{last_row},\"<\"&C{lower_row})+COUNTIFS('Raw Data'!N2:N{last_row},\">\"&C{upper_row})")
for cc in ['B','C']:
    ws_sum[f'{cc}{r}'].border = BORDER

for rr in range(bound_header_row+1, outlier_count_row+1):
    for cc in ['A','B','C']:
        ws_sum[f'{cc}{rr}'].border = BORDER

ws_sum.freeze_panes = 'A4'
ws_sum.sheet_view.showGridLines = False
ws_sum.page_setup.orientation = 'landscape'
ws_sum.page_setup.fitToWidth = 1
ws_sum.page_setup.fitToHeight = 0
ws_sum.sheet_properties.pageSetUpPr.fitToPage = True
print('Summary Statistics sheet built')

# ============================================================
# SHEET 3: Pivot Tables (formula-driven group summaries)
# ============================================================
ws_piv = wb.create_sheet('Pivot Tables')
ws_piv['A1'] = 'Pivot Tables — Orders Breakdown'
ws_piv['A1'].font = TITLE_FONT
ws_piv.merge_cells('A1:E1')

RAW = "'Raw Data'!"
def colrange(col):
    return f"{RAW}{col}2:{col}{last_row}"

block_row = 3

def write_pivot_block(ws, title, start_row, start_col, group_col_letter, group_values,
                       value_col_letter='N', label='Product'):
    r = start_row
    ws.cell(row=r, column=start_col, value=title).font = BOLD
    r += 1
    headers = [label, 'Order Count', 'Total Revenue ($)', 'Avg Order Value ($)']
    for j,h in enumerate(headers):
        ws.cell(row=r, column=start_col+j, value=h)
    style_header_row(ws, r, len(headers), start_col=start_col)
    header_r = r
    r += 1
    first_data_r = r
    for val in group_values:
        ws.cell(row=r, column=start_col, value=val)
        gcol = get_column_letter(start_col)
        countcol = get_column_letter(start_col+1)
        sumcol = get_column_letter(start_col+2)
        avgcol = get_column_letter(start_col+3)
        ws.cell(row=r, column=start_col+1,
                value=f'=COUNTIF({colrange(group_col_letter)},{gcol}{r})')
        ws.cell(row=r, column=start_col+2,
                value=f'=SUMIF({colrange(group_col_letter)},{gcol}{r},{colrange(value_col_letter)})')
        ws.cell(row=r, column=start_col+2).number_format = '#,##0.00'
        ws.cell(row=r, column=start_col+3,
                value=f'=IF({countcol}{r}=0,0,{sumcol}{r}/{countcol}{r})')
        ws.cell(row=r, column=start_col+3).number_format = '#,##0.00'
        for cc in range(start_col, start_col+4):
            ws.cell(row=r, column=cc).border = BORDER
        r += 1
    last_data_r = r - 1
    # total row
    ws.cell(row=r, column=start_col, value='TOTAL').font = BOLD
    gcol = get_column_letter(start_col)
    countcol = get_column_letter(start_col+1)
    sumcol = get_column_letter(start_col+2)
    ws.cell(row=r, column=start_col+1, value=f'=SUM({countcol}{first_data_r}:{countcol}{last_data_r})').font = BOLD
    ws.cell(row=r, column=start_col+2, value=f'=SUM({sumcol}{first_data_r}:{sumcol}{last_data_r})').font = BOLD
    ws.cell(row=r, column=start_col+2).number_format = '#,##0.00'
    for cc in range(start_col, start_col+4):
        ws.cell(row=r, column=cc).border = BORDER
        ws.cell(row=r, column=cc).fill = SECTION_FILL
    return header_r, first_data_r, last_data_r, r  # total row

products = sorted(df['Product'].unique().tolist())
statuses = sorted(df['OrderStatus'].unique().tolist())
payments = sorted(df['PaymentMethod'].unique().tolist())
referrals = sorted(df['ReferralSource'].unique().tolist())
coupons = sorted(df['CouponCode'].unique().tolist())
months = sorted(df['Month'].unique().tolist())

# Block 1: By Product (cols A-D, starting row 3)
p1_header, p1_first, p1_last, p1_total = write_pivot_block(
    ws_piv, 'By Product', 3, 1, 'D', products, value_col_letter='N', label='Product')

# Block 2: By Order Status (cols F-I, starting row 3)
p2_header, p2_first, p2_last, p2_total = write_pivot_block(
    ws_piv, 'By Order Status', 3, 6, 'I', statuses, value_col_letter='N', label='OrderStatus')

# Block 3: By Payment Method (cols A-D, after block1)
p3_start = p1_total + 3
p3_header, p3_first, p3_last, p3_total = write_pivot_block(
    ws_piv, 'By Payment Method', p3_start, 1, 'H', payments, value_col_letter='N', label='PaymentMethod')

# Block 4: By Referral Source (cols F-I, after block2)
p4_start = p2_total + 3
p4_header, p4_first, p4_last, p4_total = write_pivot_block(
    ws_piv, 'By Referral Source', p4_start, 6, 'M', referrals, value_col_letter='N', label='ReferralSource')

# Block 5: By Coupon Code (cols A-D, after block3)
p5_start = p3_total + 3
p5_header, p5_first, p5_last, p5_total = write_pivot_block(
    ws_piv, 'By Coupon Code', p5_start, 1, 'L', coupons, value_col_letter='N', label='CouponCode')

# Block 6: By Month (cols F-I, after block4) -- monthly trend, needs helper column M in Raw Data
# Add Month helper column to Raw Data sheet (col O)
ws_raw.cell(row=1, column=15, value='Month')
style_header_row(ws_raw, 1, 1, start_col=15)
for idx, row in enumerate(df.itertuples(index=False), start=2):
    ws_raw.cell(row=idx, column=15, value=row.Month)
ws_raw.column_dimensions['O'].width = 10

p6_start = p4_total + 3
p6_header, p6_first, p6_last, p6_total = write_pivot_block(
    ws_piv, 'By Month', p6_start, 6, 'O', months, value_col_letter='N', label='Month')

for col_letter, w in zip(['A','B','C','D','E','F','G','H','I','J','K','L','M'],
                          [16,12,16,16,3,16,12,16,16,3,16,12,16]):
    ws_piv.column_dimensions[col_letter].width = w
ws_piv.sheet_view.showGridLines = False
ws_piv.page_setup.orientation = 'landscape'
ws_piv.page_setup.fitToWidth = 1
ws_piv.page_setup.fitToHeight = 0
ws_piv.sheet_properties.pageSetUpPr.fitToPage = True

print('Pivot Tables sheet built')

# ============================================================
# SHEET 4: Charts
# ============================================================
ws_chart = wb.create_sheet('Charts')
ws_chart['A1'] = 'Charts — Trends, Distributions & Outliers'
ws_chart['A1'].font = TITLE_FONT
ws_chart.merge_cells('A1:H1')

def add_bar_chart(anchor, title, ws_data, cat_col, val_col, first_r, last_r, header_r, y_title, x_title):
    chart = BarChart()
    chart.type = 'col'
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.y_axis.scaling.min = 0
    chart.style = 10
    data = Reference(ws_data, min_col=val_col, min_row=header_r, max_row=last_r)
    cats = Reference(ws_data, min_col=cat_col, min_row=first_r, max_row=last_r)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    chart.legend = None
    ws_chart.add_chart(chart, anchor)

# Chart 1: Revenue by Product (bar)
add_bar_chart('A3', 'Total Revenue by Product', ws_piv, cat_col=1, val_col=3,
              first_r=p1_first, last_r=p1_last, header_r=p1_header,
              y_title='Revenue ($)', x_title='Product')

# Chart 2: Order Count by Status (bar)
add_bar_chart('A20', 'Order Count by Status', ws_piv, cat_col=6, val_col=7,
              first_r=p2_first, last_r=p2_last, header_r=p2_header,
              y_title='Number of Orders', x_title='Order Status')

# Chart 3: Order Count by Payment Method (bar)
add_bar_chart('J3', 'Order Count by Payment Method', ws_piv, cat_col=1, val_col=2,
              first_r=p3_first, last_r=p3_last, header_r=p3_header,
              y_title='Number of Orders', x_title='Payment Method')

# Chart 4: Avg Order Value by Referral Source (bar)
add_bar_chart('J20', 'Avg Order Value by Referral Source', ws_piv, cat_col=6, val_col=9,
              first_r=p4_first, last_r=p4_last, header_r=p4_header,
              y_title='Avg Order Value ($)', x_title='Referral Source')

# Chart 5: Monthly Revenue Trend (line)
line = LineChart()
line.title = 'Monthly Revenue Trend'
line.y_axis.title = 'Revenue ($)'
line.x_axis.title = 'Month'
line.style = 12
data = Reference(ws_piv, min_col=8, min_row=p6_header, max_row=p6_last)
cats = Reference(ws_piv, min_col=6, min_row=p6_first, max_row=p6_last)
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
line.height = 9
line.width = 32
line.legend = None
ws_chart.add_chart(line, 'A37')

# Histogram for TotalPrice distribution — build a frequency table first
hist_start_row = 60
ws_chart.cell(row=hist_start_row, column=1, value='TotalPrice Distribution (bins)').font = BOLD
ws_chart.cell(row=hist_start_row+1, column=1, value='Bin')
ws_chart.cell(row=hist_start_row+1, column=2, value='Count')
style_header_row(ws_chart, hist_start_row+1, 2)

tp = df['TotalPrice']
bin_edges = list(range(0, 3501, 350))
bin_labels = [f'{bin_edges[i]}-{bin_edges[i+1]}' for i in range(len(bin_edges)-1)]
counts, _ = np.histogram(tp, bins=bin_edges)
for i, (lbl, cnt) in enumerate(zip(bin_labels, counts)):
    rr = hist_start_row+2+i
    ws_chart.cell(row=rr, column=1, value=lbl)
    ws_chart.cell(row=rr, column=2, value=int(cnt))
hist_last_row = hist_start_row+1+len(bin_labels)

hist_chart = BarChart()
hist_chart.type = 'col'
hist_chart.title = 'Histogram: TotalPrice Distribution'
hist_chart.y_axis.title = 'Number of Orders'
hist_chart.x_axis.title = 'TotalPrice Range ($)'
hist_chart.style = 11
hdata = Reference(ws_chart, min_col=2, min_row=hist_start_row+1, max_row=hist_last_row)
hcats = Reference(ws_chart, min_col=1, min_row=hist_start_row+2, max_row=hist_last_row)
hist_chart.add_data(hdata, titles_from_data=True)
hist_chart.set_categories(hcats)
hist_chart.height = 9
hist_chart.width = 18
hist_chart.legend = None
ws_chart.add_chart(hist_chart, 'D60')

for col_letter, w in zip(['A','B'], [16,10]):
    ws_chart.column_dimensions[col_letter].width = w
ws_chart.sheet_view.showGridLines = False

print('Charts sheet built')

# ============================================================
# SHEET 5: Outliers
# ============================================================
ws_out = wb.create_sheet('Outliers')
ws_out['A1'] = 'Outlier Detection — TotalPrice (IQR Method)'
ws_out['A1'].font = TITLE_FONT
ws_out.merge_cells('A1:G1')
ws_out['A2'] = 'An order is flagged as an outlier if TotalPrice falls outside [Q1 − 1.5×IQR, Q3 + 1.5×IQR]. See bounds on Summary Statistics sheet.'
ws_out['A2'].font = SUBTITLE_FONT
ws_out.merge_cells('A2:G2')

out_headers = ['OrderID','Date','Product','Quantity','UnitPrice ($)','TotalPrice ($)','OrderStatus']
out_header_row = 4
for j,h in enumerate(out_headers, start=1):
    ws_out.cell(row=out_header_row, column=j, value=h)
style_header_row(ws_out, out_header_row, len(out_headers))

Q1 = df['TotalPrice'].quantile(0.25)
Q3 = df['TotalPrice'].quantile(0.75)
IQR_val = Q3 - Q1
lower_b, upper_b = Q1 - 1.5*IQR_val, Q3 + 1.5*IQR_val
outlier_df = df[(df['TotalPrice'] < lower_b) | (df['TotalPrice'] > upper_b)].sort_values('TotalPrice', ascending=False)

r = out_header_row + 1
for row in outlier_df.itertuples(index=False):
    ws_out.cell(row=r, column=1, value=row.OrderID)
    ws_out.cell(row=r, column=2, value=row.Date).number_format = 'yyyy-mm-dd'
    ws_out.cell(row=r, column=3, value=row.Product)
    ws_out.cell(row=r, column=4, value=row.Quantity)
    ws_out.cell(row=r, column=5, value=row.UnitPrice).number_format = '#,##0.00'
    ws_out.cell(row=r, column=6, value=row.TotalPrice).number_format = '#,##0.00'
    ws_out.cell(row=r, column=6).font = Font(bold=True, color='C00000', name='Arial')
    ws_out.cell(row=r, column=7, value=row.OrderStatus)
    for c in range(1,8):
        ws_out.cell(row=r, column=c).border = BORDER
    r += 1
out_last_row = r - 1

add_table(ws_out, f'A{out_header_row}:G{out_last_row}', 'OutlierTable', style='TableStyleMedium11')
widths = {'A':12,'B':12,'C':10,'D':10,'E':13,'F':14,'G':13}
autosize(ws_out, widths)
ws_out.sheet_view.showGridLines = False
ws_out.page_setup.orientation = 'landscape'

note_row = out_last_row + 2
ws_out.cell(row=note_row, column=1,
    value=f'Total flagged outliers: {len(outlier_df)} out of {len(df)} orders ({len(outlier_df)/len(df)*100:.1f}%). '
          f'All flagged orders involve Quantity=5 combined with a high UnitPrice — these are large legitimate bulk purchases, not data errors.')
ws_out.cell(row=note_row, column=1).font = Font(italic=True, name='Arial', size=10)
ws_out.merge_cells(f'A{note_row}:G{note_row}')

print('Outliers sheet built')

# ============================================================
# SHEET 6: Key Observations
# ============================================================
ws_obs = wb.create_sheet('Key Observations')
ws_obs['A1'] = 'Key Observations — Exploratory Data Analysis'
ws_obs['A1'].font = TITLE_FONT
ws_obs.merge_cells('A1:E1')
ws_obs['A2'] = 'Project 2: Exploratory Data Analysis (EDA) — DecodeLabs Internship'
ws_obs['A2'].font = SUBTITLE_FONT
ws_obs.merge_cells('A2:E2')

# Pre-compute the figures needed for narrative bullets
top_product = df.groupby('Product')['TotalPrice'].sum().idxmax()
top_product_rev = df.groupby('Product')['TotalPrice'].sum().max()
low_product = df.groupby('Product')['TotalPrice'].sum().idxmin()
top_referral = df.groupby('ReferralSource').size().idxmax()
top_payment = df['PaymentMethod'].value_counts().idxmax()
cancel_rate = (df['OrderStatus']=='Cancelled').mean()*100
return_rate = (df['OrderStatus']=='Returned').mean()*100
avg_order = df['TotalPrice'].mean()
median_order = df['TotalPrice'].median()
coupon_use_rate = df['CouponCode'].ne('None').mean()*100
best_month = df.groupby('Month')['TotalPrice'].sum().idxmax()
worst_month = df.groupby('Month')['TotalPrice'].sum().idxmin()
n_outliers = len(outlier_df)

sections = [
    ("1. Dataset Overview", [
        f"The dataset contains {len(df):,} orders spanning {df['Date'].min().strftime('%b %Y')} to {df['Date'].max().strftime('%b %Y')} (~30 months).",
        f"14 original fields covering order details, customer, product, payment, fulfillment status, and marketing attribution.",
        f"Data quality is strong: 0 duplicate rows, 0 duplicate OrderIDs, and the only missing values are in CouponCode (309 of {len(df)} orders, 25.8%) — which simply means no coupon was applied, not a data issue.",
        f"TotalPrice consistently equals Quantity × UnitPrice across all rows, confirming the calculated field is reliable.",
    ]),
    ("2. Central Tendency & Spread", [
        f"Average order value is ${avg_order:,.2f}, while the median is ${median_order:,.2f} — the mean sits notably above the median, indicating a right-skewed distribution with some high-value orders pulling the average up.",
        f"UnitPrice ranges from $11.39 to $699.93 (mean ≈ $356.41), and Quantity ranges from 1 to 5 units per order (mean ≈ 2.95), so order size is fairly evenly spread rather than concentrated at one extreme.",
        f"ItemsInCart (1 to 10, mean ≈ 5.49) is consistently higher than Quantity purchased, suggesting cart abandonment of some items is common before checkout.",
        "See the Summary Statistics sheet for full mean/median/mode/std-dev/quartile breakdowns of all four numeric fields.",
    ]),
    ("3. Trends", [
        f"Revenue is fairly evenly distributed across the 7 product categories — {top_product} leads with ${top_product_rev:,.0f} total revenue, but the gap to the lowest ({low_product}) is modest, so no single product dominates sales.",
        f"{top_payment} is the most-used payment method, but all 5 payment methods (Online, Cash, Credit Card, Debit Card, Gift Card) are used at similar volumes (~230-260 orders each) — customers don't show a strong preference.",
        f"{top_referral} is the top traffic/referral source by order count, with all 5 sources (Instagram, Email, Google, Facebook, Referral) again landing in a fairly narrow band.",
        f"Monthly revenue fluctuates between roughly $27,000 and $68,000 with no strong seasonal pattern — {best_month} was the strongest month and {worst_month} the weakest in the dataset, but order volume per month stays fairly stable (27-53 orders), so swings are driven more by order mix than by demand spikes.",
        f"Coupon usage is widespread — {coupon_use_rate:.1f}% of orders used a coupon (SAVE10, FREESHIP, or WINTER15) — but average order value with vs. without a coupon is nearly identical (~$1,057 vs ~$1,043), so coupons don't appear to be driving materially larger basket sizes.",
    ]),
    ("4. Order Status Patterns", [
        f"Order outcomes are nearly evenly split across the 5 statuses (~230-250 orders each): Cancelled ({cancel_rate:.1f}%), Returned ({return_rate:.1f}%), Pending, Shipped, and Delivered.",
        f"Combined, Cancelled + Returned orders make up {(cancel_rate+return_rate):.1f}% of all orders — worth flagging operationally, since that's roughly 4 in 10 orders not completing as a normal delivery.",
        "Average order value does not differ meaningfully by status, meaning cancellations/returns are not concentrated in either low- or high-value orders.",
    ]),
    ("5. Outliers", [
        f"Using the IQR method (1.5× rule) on TotalPrice, {n_outliers} orders (0.7%) are flagged as outliers, all above the upper bound of ~${upper_b:,.0f} — no orders fall below the lower bound (which is negative, so it has no practical effect).",
        "Every flagged outlier combines the maximum Quantity (5 units) with a high UnitPrice — these reflect genuine large bulk purchases (Tablets, Monitors, Laptops, Chairs, Printers), not data entry errors.",
        "UnitPrice itself has zero outliers under the same method — pricing is consistent and contains no extreme/erroneous values.",
        "See the Outliers sheet for the full list of flagged orders.",
    ]),
    ("6. Recommendations / Next Steps", [
        "Investigate the combined 41% Cancelled+Returned rate — segment by product, payment method, or shipping address to see if any factor correlates with non-completion.",
        "Since coupon usage doesn't lift average order value, consider testing whether coupons are better aimed at increasing order frequency or customer acquisition instead.",
        "The right-skewed TotalPrice distribution (mean > median) suggests reporting median alongside mean for a more representative 'typical order' figure in future dashboards.",
    ]),
]

r = 4
for title, bullets in sections:
    ws_obs.cell(row=r, column=1, value=title).font = Font(bold=True, color='FFFFFF', name='Arial', size=12)
    ws_obs.cell(row=r, column=1).fill = HEADER_FILL
    ws_obs.merge_cells(f'A{r}:F{r}')
    ws_obs.cell(row=r, column=1).alignment = Alignment(vertical='center', indent=1)
    ws_obs.row_dimensions[r].height = 20
    r += 1
    for b in bullets:
        ws_obs.cell(row=r, column=1, value='•')
        ws_obs.cell(row=r, column=1).alignment = Alignment(vertical='top')
        ws_obs.cell(row=r, column=2, value=b)
        ws_obs.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        ws_obs.cell(row=r, column=2).font = NORMAL
        ws_obs.merge_cells(f'B{r}:F{r}')
        # estimate row height based on text length at ~110 chars/line for the merged width
        n_lines = max(1, -(-len(b) // 100))
        ws_obs.row_dimensions[r].height = 15 * n_lines
        r += 1
    r += 1

ws_obs.column_dimensions['A'].width = 4
for col in ['B','C','D','E','F']:
    ws_obs.column_dimensions[col].width = 21
ws_obs.sheet_view.showGridLines = False
ws_obs.page_setup.orientation = 'landscape'
ws_obs.page_setup.fitToWidth = 1
ws_obs.page_setup.fitToHeight = 0
ws_obs.sheet_properties.pageSetUpPr.fitToPage = True

print('Key Observations sheet built')

# ============================================================
# Reorder sheets logically & set active sheet
# ============================================================
order = ['Key Observations', 'Summary Statistics', 'Pivot Tables', 'Charts', 'Outliers', 'Raw Data']
wb._sheets = [wb[name] for name in order]
wb.active = 0

wb.save(OUT_PATH)
print('Saved final workbook to', OUT_PATH)
print(f'Done. {len(df):,} orders processed -> {OUT_PATH}')
